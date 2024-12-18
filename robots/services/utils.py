from datetime import datetime, timedelta

from django.db.models import Count
from django.http import QueryDict, FileResponse
from django.utils import timezone
from openpyxl import Workbook

from robots.models import Robot
from robots.services import RobotDDO


def validate_data_robot(data: QueryDict) -> RobotDDO:
    model = data.get("model", None)
    version = data.get("version", None)
    created = data.get("created", None)
    return RobotDDO(
        model=model,
        version=version,
        created=created,
    )


def save_new_robot(robot: RobotDDO) -> None:
    Robot.objects.create(
        serial=robot.serial,
        model=robot.model,
        version=robot.version,
        created=robot.created
    )

def get_report_period(period: int = 7) -> tuple[datetime, datetime]:
    end_date = timezone.now()
    start_date = end_date - timedelta(days=period)
    return start_date, end_date


def get_model_version_count_robots(start_date: datetime, end_date: datetime):
    return (
        Robot.objects
        .filter(created__range=[start_date, end_date])
        .values("model", "version")
        .annotate(count=Count("id"))
        .order_by("model", "version")
    )


def create_new_robot_report(start_date: datetime, end_date: datetime) -> str:
    robots = get_model_version_count_robots(start_date, end_date)
    path_file = fr"robots/data/robot_reports/robots_report_{start_date.strftime('%d.%m.%Y')}-{end_date.strftime('%d.%m.%Y')}.xlsx"
    wb = Workbook()
    for robot in robots:
        if robot["model"] not in wb.sheetnames:
            wb.create_sheet(f"{robot['model']}")
        sheet = wb[f"{robot['model']}"]
        sheet.append([robot["model"], robot["version"], robot["count"]])
    wb.remove_sheet(wb["Sheet"])
    wb.save(path_file)
    return path_file


def create_file_response(path_to_file: str, start_date: datetime, end_date: datetime) -> FileResponse:
    file = open(path_to_file, 'rb')
    response = FileResponse(file, content_type='application/octet-stream')
    response['Content-Disposition'] = (f'attachment; filename="robots report {start_date.strftime("%d.%m.%Y")}'
                                       f'-{end_date.strftime("%d.%m.%Y")}.xlsx"')
    return response

def get_robots_by_serial(serial: str) -> list[Robot]:
    return Robot.objects.filter(serial=serial)
